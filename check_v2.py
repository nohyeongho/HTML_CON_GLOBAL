import openpyxl

# V2 파일 구조 확인
xlsx_file = "global_model V2.xlsx"
wb = openpyxl.load_workbook(xlsx_file, data_only=True)
sheet = wb.active
print('Sheet name:', sheet.title)
print('Max row:', sheet.max_row)
print('Max col:', sheet.max_column)
print('--- 첫 5행 미리보기 ---')
for i, row in enumerate(sheet.iter_rows(min_row=1, max_row=5, values_only=True)):
    print(f'Row {i+1}:', row)
